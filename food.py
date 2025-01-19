from openpyxl import load_workbook

class Food:
    def __init__(self, excel_path):
        self.path = excel_path

    def read_excel(self):
        food_dict = {}
        wb = load_workbook(self.path)
        food_type = wb.sheetnames
        for type in food_type:
            food_dict[type] = {}
            ws = wb[type]
            for row in range(ws.max_row):
                row_index = row + 1
                name = ws.cell(row_index, 1).value
                price = ws.cell(row_index, 2).value
                food_dict[type][name] = price
        
        return food_dict
